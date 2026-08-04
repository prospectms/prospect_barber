"""Categoria 1 — nenhuma rota autenticada da empresa A retorna dado da
empresa B: listagens, dashboard e relatórios."""
import io
import openpyxl


def _login_dono_na_unidade(client, login, empresa):
    login(empresa["dono_email"], empresa["senha"])
    client.post("/auth/unidade", data={"unidade_id": empresa["unidade_id"]}, follow_redirects=True)


def test_customers_index_nao_vaza_cliente_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/customers/")
    assert r.status_code == 200
    assert empresa_a["cliente_nome"] in r.text
    assert empresa_b["cliente_nome"] not in r.text


def test_barbers_index_nao_vaza_profissional_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/barbers/")
    assert r.status_code == 200
    assert empresa_a["prof_nome"] in r.text
    assert empresa_b["prof_nome"] not in r.text


def test_services_index_nao_vaza_servico_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/services/")
    assert r.status_code == 200
    assert empresa_a["servico_nome"] in r.text
    assert empresa_b["servico_nome"] not in r.text


def test_appointments_index_nao_vaza_agendamento_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/appointments/")
    assert r.status_code == 200
    assert empresa_a["cliente_nome"] in r.text
    assert empresa_b["cliente_nome"] not in r.text
    assert empresa_b["prof_nome"] not in r.text


def test_auth_users_nao_vaza_usuario_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/auth/users")
    assert r.status_code == 200
    assert empresa_a["dono_email"] in r.text
    assert empresa_b["dono_email"] not in r.text
    assert empresa_b["func_email"] not in r.text


def test_dashboard_stats_nao_conta_cliente_de_outra_empresa(client, login, empresa_a, empresa_b):
    """total_customers no dashboard é a contagem de Cliente.query.count() —
    se vazasse, contaria 2 (1 de cada empresa) em vez de 1."""
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/")
    assert r.status_code == 200
    assert empresa_a["prof_nome"] in r.text  # top_barbers / upcoming devem mostrar só o proprio
    assert empresa_b["prof_nome"] not in r.text
    assert empresa_b["cliente_nome"] not in r.text


def test_reports_export_xlsx_nao_vaza_dado_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/reports/export?report=appointments&fmt=xlsx")
    assert r.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    all_text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )
    assert empresa_a["cliente_nome"] in all_text
    assert empresa_b["cliente_nome"] not in all_text
    assert empresa_b["prof_nome"] not in all_text


def test_reports_export_xlsx_barbeiros_nao_vaza_de_outra_empresa(client, login, empresa_a, empresa_b):
    _login_dono_na_unidade(client, login, empresa_a)
    r = client.get("/reports/export?report=barbers&fmt=xlsx")
    assert r.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    ws = wb.active
    all_text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )
    assert empresa_b["prof_nome"] not in all_text
